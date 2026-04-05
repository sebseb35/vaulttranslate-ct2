from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook

from packages.core import DocumentAdapter, DocumentFormat, Segment


@dataclass(frozen=True, slots=True)
class _TextTarget:
    cell: Cell


class XlsxDocumentAdapter(DocumentAdapter):
    @property
    def supported_format(self) -> DocumentFormat:
        return DocumentFormat.XLSX

    def extract_segments(self, document_bytes: bytes) -> tuple[Segment, ...]:
        workbook = _load_xlsx(document_bytes)
        targets = _collect_text_targets(workbook)
        return tuple(
            Segment(
                segment_id=f"xlsx-{index}",
                text=target.cell.value,
            )
            for index, target in enumerate(targets)
        )

    def rebuild_document(
        self, original_document: bytes, translated_segments: Sequence[Segment]
    ) -> bytes:
        workbook = _load_xlsx(original_document)
        targets = _collect_text_targets(workbook)

        if len(translated_segments) != len(targets):
            raise ValueError("XLSX rebuild requires the same number of segments as extraction")

        for index, (target, translated_segment) in enumerate(
            zip(targets, translated_segments, strict=True)
        ):
            expected_id = f"xlsx-{index}"
            if translated_segment.segment_id != expected_id:
                raise ValueError(
                    "XLSX segment mismatch: "
                    f"expected '{expected_id}' but got '{translated_segment.segment_id}'"
                )
            target.cell.value = translated_segment.text

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


def _load_xlsx(document_bytes: bytes) -> Workbook:
    try:
        return load_workbook(BytesIO(document_bytes))
    except Exception as exc:  # pragma: no cover - defensive handling
        raise ValueError("Failed to parse XLSX document bytes") from exc


def _collect_text_targets(workbook: Workbook) -> list[_TextTarget]:
    targets: list[_TextTarget] = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if _is_translatable_text_cell(cell):
                    targets.append(_TextTarget(cell=cell))

    return targets


def _is_translatable_text_cell(cell: Cell) -> bool:
    value = cell.value
    if not isinstance(value, str):
        return False
    if value == "":
        return False
    if cell.data_type == "f":
        return False
    return True
