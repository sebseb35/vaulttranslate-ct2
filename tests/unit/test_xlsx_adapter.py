from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from packages.core.models import DocumentFormat, Segment


FIXTURES_DIR = Path(__file__).resolve().parents[1] / "fixtures" / "xlsx"


def _fixture_bytes(name: str) -> bytes:
    return (FIXTURES_DIR / name).read_bytes()


def test_xlsx_adapter_extracts_only_string_cells() -> None:
    from packages.adapter_xlsx import XlsxDocumentAdapter

    adapter = XlsxDocumentAdapter()
    source = _fixture_bytes("sample.xlsx")

    segments = adapter.extract_segments(source)

    assert adapter.supported_format is DocumentFormat.XLSX
    assert segments == (
        Segment(segment_id="xlsx-0", text="Hello"),
        Segment(segment_id="xlsx-1", text="World"),
        Segment(segment_id="xlsx-2", text="Notes"),
        Segment(segment_id="xlsx-3", text="Final value"),
    )


def test_xlsx_adapter_extracts_edge_case_string_cells_in_sheet_order() -> None:
    from packages.adapter_xlsx import XlsxDocumentAdapter

    adapter = XlsxDocumentAdapter()
    source = _fixture_bytes("edge_cases.xlsx")

    segments = adapter.extract_segments(source)

    assert [segment.text for segment in segments] == [
        "  padded text  ",
        "   ",
        "Merged title",
        "Line 1\nLine 2",
        "Résumé",
        "SKU",
        "Description",
        "A-100",
        "Widget basic",
    ]


def test_xlsx_adapter_rebuild_updates_text_cells_and_preserves_structure() -> None:
    from packages.adapter_xlsx import XlsxDocumentAdapter

    adapter = XlsxDocumentAdapter()
    source = _fixture_bytes("sample.xlsx")
    extracted = adapter.extract_segments(source)

    translated = tuple(
        Segment(segment_id=segment.segment_id, text=f"{segment.text} FR")
        for segment in extracted
    )

    rebuilt_bytes = adapter.rebuild_document(source, translated)
    rebuilt = load_workbook(BytesIO(rebuilt_bytes))

    assert rebuilt.sheetnames == ["Summary", "Details"]
    assert rebuilt["Summary"]["A1"].value == "Hello FR"
    assert rebuilt["Summary"]["A2"].value == "World FR"
    assert rebuilt["Summary"]["B1"].value == 42
    assert rebuilt["Summary"]["C1"].value == "=SUM(B1,8)"
    assert rebuilt["Details"]["A1"].value == "Notes FR"
    assert rebuilt["Details"]["B3"].value == "Final value FR"


def test_xlsx_adapter_rebuild_preserves_formulas_merged_cells_and_non_string_values() -> None:
    from packages.adapter_xlsx import XlsxDocumentAdapter

    adapter = XlsxDocumentAdapter()
    source = _fixture_bytes("edge_cases.xlsx")
    extracted = adapter.extract_segments(source)

    translated = tuple(
        Segment(segment_id=segment.segment_id, text=f"[fr] {segment.text}")
        for segment in extracted
    )

    rebuilt_bytes = adapter.rebuild_document(source, translated)
    rebuilt = load_workbook(BytesIO(rebuilt_bytes))

    edge_sheet = rebuilt["EdgeCases"]
    inventory_sheet = rebuilt["Inventory"]

    assert edge_sheet["A1"].value == "[fr]   padded text  "
    assert edge_sheet["B1"].value == "[fr]    "
    assert edge_sheet["C1"].value is True
    assert edge_sheet["D1"].value.year == 2024
    assert edge_sheet["E1"].value == '=CONCAT("prefix",A1)'
    assert "A3:B3" in {str(range_ref) for range_ref in edge_sheet.merged_cells.ranges}
    assert edge_sheet["A3"].value == "[fr] Merged title"
    assert edge_sheet["B3"].value is None
    assert edge_sheet["A5"].value == "[fr] Line 1\nLine 2"
    assert edge_sheet["B5"].value == 0
    assert edge_sheet["C5"].value == "[fr] Résumé"
    assert inventory_sheet["A1"].value == "[fr] SKU"
    assert inventory_sheet["B1"].value == "[fr] Description"
    assert inventory_sheet["A2"].value == "[fr] A-100"
    assert inventory_sheet["B2"].value == "[fr] Widget basic"
    assert inventory_sheet["C2"].value == '=A2&"-FR"'


def test_xlsx_adapter_rebuild_validates_segment_count_and_ids() -> None:
    from packages.adapter_xlsx import XlsxDocumentAdapter

    adapter = XlsxDocumentAdapter()
    source = _fixture_bytes("sample.xlsx")
    extracted = adapter.extract_segments(source)

    with pytest.raises(ValueError, match="same number of segments"):
        adapter.rebuild_document(source, extracted[:-1])

    wrong_first = (Segment(segment_id="xlsx-9999", text="bad"),) + extracted[1:]
    with pytest.raises(ValueError, match="segment mismatch"):
        adapter.rebuild_document(source, wrong_first)
